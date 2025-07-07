import pdfplumber
import pandas as pd
import re
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment

# ------------------------------------------
# ------------------------------------------
# File paths - FIXED OUTPUT DIR (REMOVED TRAILING SPACE)
pdf_path = "D:/UNIVERSITE D'AIX MARSEILLE/Day 1 03_03_2025/X2R1/X2R1-T5.3-D-SIE-102-20_-_D5.1_-_Moving_Block_System_Requirements.pdf"
pdf_filename = os.path.basename(pdf_path).replace(".pdf", "")
output_dir = "C:/Users/aroua/Desktop/correction"  # No trailing space

# Ensure output directory exists - CRITICAL FIX
os.makedirs(output_dir, exist_ok=True)
output_excel = os.path.join(output_dir, f"{pdf_filename}_Result.xlsx")

# Create Excel workbook
wb = Workbook()
ws = wb.active
ws.title = "Extracted Requirements"

# Headers
required_columns = ["Topic", "Requirement ID", "Description", "Traceability"]
for col_idx, header in enumerate(required_columns):
    ws.cell(row=1, column=col_idx + 1, value=header)

# Regex patterns
req_id_pattern = re.compile(r"REQ-[A-Za-z0-9]+-\d+")  # Plus flexible
traceability_pattern = re.compile(r"\[(X2R\d+ D\d+\.\d+: REQ-[A-Za-z0-9-]+)\]")
footer_pattern = re.compile(r"(GA\s*\d+\s*)?Page\s+\d+\s+of\s+\d+", re.IGNORECASE)
block_delimiter_pattern = re.compile(r"(Requirement:|Rationale:|Guidance:|Introduction)", re.IGNORECASE)

def is_valid_requirement_block(lines, idx):
    """Vérifie si l'ID est dans un bloc 'Requirement' valide"""
    # Vérifier 3 lignes avant et 2 lignes après
    start = max(0, idx - 3)
    end = min(len(lines), idx + 2)
    
    for i in range(start, end):
        if re.search(r"^\s*Requirement:", lines[i], re.IGNORECASE):
            return True
    return False

def extract_description(lines, start_idx):
    """Extrait la description jusqu'au prochain délimiteur"""
    description = []
    for i in range(start_idx, len(lines)):
        line = footer_pattern.sub("", lines[i]).strip()
        
        # Arrêt aux délimiteurs de section
        if block_delimiter_pattern.search(line) and i > start_idx:
            break
            
        if line:
            description.append(line)
    return "\n".join(description).strip()

def extract_requirements(pdf_path):
    requirements = []
    current_topic = "Unknown"
    last_traceability = "[Not Provided]"
    in_requirement_block = False

    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            text = page.extract_text(layout=True)
            if not text:
                continue

            lines = text.split("\n")
            for idx, line in enumerate(lines):
                # Clean line first
                clean_line = footer_pattern.sub("", line).strip()
                if not clean_line:
                    continue

                # 1. Détection des sujets
                topic_match = re.match(r"^\s*(\d+\.\d+)\s+([A-Za-z][A-Za-z0-9 \-/]+$)", clean_line)
                if topic_match:
                    current_topic = topic_match.group(2).strip()
                    in_requirement_block = False
                    continue

                # 2. Détection de la traçabilité
                trace_match = traceability_pattern.search(clean_line)
                if trace_match:
                    last_traceability = trace_match.group(1).strip()
                elif "[New]" in clean_line:
                    last_traceability = "New"

                # 3. Détection des délimiteurs de sections
                if block_delimiter_pattern.search(clean_line):
                    in_requirement_block = "Requirement:" in clean_line
                    continue

                # 4. Détection des exigences VALIDES
                req_id_match = req_id_pattern.search(clean_line)
                if req_id_match and in_requirement_block:
                    req_id = req_id_match.group().strip()
                    
                    # Validation supplémentaire
                    if not is_valid_requirement_block(lines, idx):
                        continue
                        
                    description = extract_description(lines, idx + 1)
                    requirements.append((current_topic, req_id, description, last_traceability))
                    
    return requirements

# ... (le reste du code reste inchangé) ...

# Extract and clean
data = extract_requirements(pdf_path)
extracted_df = pd.DataFrame(data, columns=required_columns)
extracted_df.drop_duplicates(subset=["Requirement ID"], keep="first", inplace=True)

# Write to Excel
for row_idx, row in extracted_df.iterrows():
    for col_idx, col_name in enumerate(required_columns):
        ws.cell(row=row_idx + 2, column=col_idx + 1, value=row[col_name])

# Formatting
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(required_columns)):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True)

# Row height for description
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3):
    for cell in row:
        ws.row_dimensions[cell.row].height = None

# Column width
for col in ws.columns:
    max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
    col_letter = col[0].column_letter
    ws.column_dimensions[col_letter].width = min(max_length + 5, 50)

# Save
wb.save(output_excel)
print(f"✅ Extraction completed! {len(extracted_df)} requirements saved to: {output_excel}")
