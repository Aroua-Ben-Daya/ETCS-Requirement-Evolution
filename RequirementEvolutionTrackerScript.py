import openpyxl  # Pour lire/écrire des fichiers Excel
from openpyxl.styles import PatternFill, Alignment  # Pour colorer et aligner les cellules
from openpyxl.utils import get_column_letter  # Pour convertir un numéro de colonne en lettre (1 -> A)
import os
import re  # Pour détecter les traçabilités avec des expressions régulières

# Liste des versions à analyser
VERSIONS = ["X2R1", "X2R3", "X2R5"]

# Icônes utilisées pour représenter les statuts des exigences
STATUS_ICONS = {
    "New": "🆕 New",
    "Unchanged": "✅ Unchanged",
    "Modified": "📝 Modified",
    "Absent": "❌ Absent"
}

# Couleurs utilisées pour chaque version (pour le fichier Excel final)
VERSION_COLORS = {
    "X2R1": "FFDAB9",   # Pêche
    "X2R3": "FFFFE0",   # Jaune clair
    "X2R5": "E0FFFF"    # Cyan clair
}

# Lecture d'un fichier Excel et extraction des exigences dans un dictionnaire
def parse_excel(file_path, version):
    from openpyxl import load_workbook
    wb = load_workbook(file_path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    data = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        topic, req_id, desc, trace = row
        if not req_id:
            continue
        data[req_id] = {
            "version": version,
            "id": req_id,
            "topic": topic.strip() if topic else "",
            "description": desc.strip() if desc else "",
            "traceability": trace.strip() if trace else ""
        }
    return data

# Extraction des références de traçabilité depuis un champ texte
def extract_trace_ids(trace_str):
    pattern = r"(X2R1|X2R3|X2R5)[^:]*:\s*(REQ-[\w\-/]+)"
    return re.findall(pattern, trace_str or "")

# Construction des chaînes de traçabilité complètes
def build_traceability_chains(requirements_by_version):
    chains = []
    seen = set()

    for version in VERSIONS:
        for req_id, req in requirements_by_version[version].items():
            key = (version, req_id)
            if key in seen:
                continue
            chain = [(version, req_id)]
            seen.add(key)

            trace = req["traceability"]
            while trace:
                matches = extract_trace_ids(trace)
                if not matches:
                    break
                prev_version, prev_req_id = matches[0]
                if prev_req_id in requirements_by_version.get(prev_version, {}):
                    chain.insert(0, (prev_version, prev_req_id))
                    seen.add((prev_version, prev_req_id))
                    trace = requirements_by_version[prev_version][prev_req_id]["traceability"]
                else:
                    break
            chains.append(chain)
    return chains

# Détermine le statut d'une exigence par rapport à sa base
def classify(curr, base):
    if not base:
        return "New", ""
    elif curr["description"] == base["description"] and curr["topic"] == base["topic"]:
        return "Unchanged", ""
    else:
        return "Modified", describe_diff(curr, base)

# Décrit les différences entre deux exigences
def describe_diff(curr, base):
    changes = []
    if curr["topic"] != base["topic"]:
        changes.append("Topic changed")
    if curr["description"] != base["description"]:
        changes.append("Description changed")
    return ", ".join(changes)

# Marque comme "Absent" les exigences qui n'ont pas été reprises dans la version suivante
def detect_absent(requirements_by_version, output):
    versions = list(requirements_by_version.keys())
    for i in range(len(versions) - 1):
        current_version = versions[i]
        next_version = versions[i + 1]

        base_ids_next_version = set(
            row["base"] for row in output if row["version"] == next_version and row["base"]
        )

        for row in output:
            if row["version"] == current_version and row["status"] == "New":
                actual_id = row["id"]
                if actual_id not in base_ids_next_version:
                    row["status"] = "Absent"
                    row["path"] += f" → {next_version}:❌"
    return output

# Génère la liste finale des exigences avec leur statut d'évolution
def generate_output(chains, requirements_by_version):
    output = []
    seen_requirements = set()
    new_candidates_to_skip = set()

    trace_refs = {}
    for version in VERSIONS:
        trace_refs[version] = set()
        for req in requirements_by_version[version].values():
            for ref_version, ref_id in extract_trace_ids(req["traceability"]):
                trace_refs[ref_version].add(ref_id)

    for chain in chains:
        for i, (version, rid) in enumerate(chain):
            req = requirements_by_version[version][rid]
            base = None
            if i > 0:
                base_version, base_rid = chain[i - 1]
                base = requirements_by_version[base_version][base_rid]
                base_id = base_rid
            else:
                base_id = ""

            status, change = classify(req, base)

            if status == "New":
                next_version_index = VERSIONS.index(version) + 1
                if next_version_index < len(VERSIONS):
                    next_version = VERSIONS[next_version_index]
                    if rid in trace_refs.get(version, set()):
                        new_candidates_to_skip.add((version, rid))
                        continue

            if (version, rid) in seen_requirements or (version, rid) in new_candidates_to_skip:
                continue

            path = " → ".join([f"{v}:{r}" for (v, r) in chain[:i + 1]])
            seen_requirements.add((version, rid))

            output.append({
                "version": version,
                "id": rid,
                "base": base_id,
                "status": status,
                "path": path,
                "change": change
            })

    output = detect_absent(requirements_by_version, output)
    return output

# Mise en forme et export des résultats dans un fichier Excel
def format_output_to_excel(output, file_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Evolution"

    headers = ["Version", "Current Requirement ID", "Base Requirement ID", "Status", "Traceability Path", "Topic or description change"]
    ws.append(headers)

    for row in output:
        excel_row = [
            row["version"],
            row["id"],
            row["base"],
            STATUS_ICONS[row["status"]],
            row["path"],
            row["change"]
        ]
        ws.append(excel_row)

        fill = PatternFill(start_color=VERSION_COLORS[row["version"]], end_color=VERSION_COLORS[row["version"]], fill_type="solid")
        for col in range(1, len(headers) + 1):
            ws.cell(row=ws.max_row, column=col).fill = fill
            ws.cell(row=ws.max_row, column=col).alignment = Alignment(wrap_text=True, vertical="top")

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 30

    wb.save(file_path)

# Point d'entrée principal du script
if __name__ == "__main__":
    input_files = {
        "X2R1": "C:/Users/aroua/Desktop/ReqEvolutionTracker/X2R1-T5.3-D-SIE-102-20_-_D5.1_-_Moving_Block_System_Requirements_Result.xlsx",
        "X2R3": "C:/Users/aroua/Desktop/ReqEvolutionTracker/X2R3-T4_3-D-SMD-008-19_-_D4.2Part3-SystemSpecification_Result.xlsx",
        "X2R5": "C:/Users/aroua/Desktop/ReqEvolutionTracker/X2R5-T4_2-D-SMD-003-23_-_D41Part3SystemSpecification_Result.xlsx"
    }

    requirements_by_version = {
        version: parse_excel(path, version)
        for version, path in input_files.items()
    }

    chains = build_traceability_chains(requirements_by_version)
    output = generate_output(chains, requirements_by_version)
    format_output_to_excel(output, "C:/Users/aroua/Desktop/RequirementEvolutionOutputDetect_absent_Changed.xlsx")
